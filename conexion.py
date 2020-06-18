import pymysql
from openpyxl import load_workbook
#En esta parte se realiza la conexion de python hacia excel, en file_path se coloca la ruta de la hoja de calculo
#y en sheet el nombre de la hoja de la cul se quieren obtener los archivos
FILE_PATH = "[RUTA DE ARCHIVO DE EXCEL]"
SHEET = '[NOMBRE DE HOJA DE CALCULO]'
def main(): 
    wb = load_workbook(FILE_PATH, read_only=True)
    sheet = wb[SHEET]
    columnname = []
    #se ve hermoso tu codigo
    #Row election tiene como objetivo seleccionar apartir de que fila se obtendran datos
    row_election = input("Select the Excel row from which you want to start uploading data: ")
    #Ciclo el cual va recorriendo los datos, el cada iterador es una columna de excel, se pueden
    #agregar tantas nombrecolumna como columnas existan en la hoja de calculo
    for row in sheet.iter_rows(min_row=row_election):
        columnname.append(row[0].value) 




    con = pymysql.connect('[AQUI SE COLOCA LA IP DE LA INSTANCIA]', '[USER]', 
    '[PASSWORD]', '[DATABASE NAME]')  
    iterObject = 0
    for row in columnname:
        try:
            with con:
                cur = con.cursor()
                #Comando el cual tiene como objetivo insertar datos en una tabla, previamente creada
                cur.execute("INSERT INTO `[Database name]`.`[Table name]` (`[Sql column name]`) VALUES ('{}');".format(columnname[iterObject]))
                rows = cur.fetchall()
                
                iterObject += 1
                continue
        except KeyboardInterrupt:
            print("exit...")
            exit()
        except:
            print("Error")




if __name__ == "__main__":
    main()



























